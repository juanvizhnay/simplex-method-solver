from openpyxl import Workbook
from openpyxl.styles import Font

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
)

from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter


def export_excel(filename, solver, result_text):
    wb = Workbook()

    ws = wb.active
    ws.title = "Resultados"

    bold = Font(bold=True)

    ws["A1"] = "SOLVER DE PROGRAMACIÓN LINEAL"
    ws["A1"].font = bold

    ws["A3"] = "Función objetivo:"
    ws["B3"] = solver.objective

    ws["A4"] = "Tipo:"
    ws["B4"] = solver.sense

    ws["A6"] = "Restricciones:"
    ws["A6"].font = bold

    row = 7

    for c in solver.constraints:
        ws[f"A{row}"] = c
        row += 1

    row += 2

    ws[f"A{row}"] = "Resultados"
    ws[f"A{row}"].font = bold

    row += 1

    for line in result_text.split("\n"):
        ws[f"A{row}"] = line
        row += 1

    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 30

    wb.save(filename)


def export_pdf(filename, solver, result_text):
    doc = SimpleDocTemplate(
        filename,
        pagesize=letter,
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "<b>SOLVER DE PROGRAMACIÓN LINEAL</b>",
            styles["Title"],
        )
    )

    elements.append(Spacer(1, 15))

    elements.append(
        Paragraph(
            f"<b>Función objetivo:</b> {solver.objective}",
            styles["BodyText"],
        )
    )

    elements.append(
        Paragraph(
            f"<b>Tipo:</b> {solver.sense}",
            styles["BodyText"],
        )
    )

    elements.append(Spacer(1, 10))

    elements.append(
        Paragraph(
            "<b>Restricciones:</b>",
            styles["Heading2"],
        )
    )

    for c in solver.constraints:
        elements.append(
            Paragraph(c, styles["BodyText"])
        )

    elements.append(Spacer(1, 10))

    elements.append(
        Paragraph(
            "<b>Resultados:</b>",
            styles["Heading2"],
        )
    )

    for line in result_text.split("\n"):
        line = line.replace(" ", "&nbsp;")

        elements.append(
            Paragraph(
                line,
                styles["Code"],
            )
        )

    doc.build(elements)